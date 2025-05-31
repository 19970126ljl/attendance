import openpyxl
import os
import argparse

def inspect_excel_file(file_path):
    print(f"\n--- Inspecting: {file_path} ---")
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    try:
        workbook = openpyxl.load_workbook(file_path)
        print(f"Sheet names: {workbook.sheetnames}")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n  --- Sheet: {sheet_name} ---")
            print("  First 5 rows (including headers):")
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5)):
                row_values = [cell.value for cell in row]
                print(f"    Row {row_idx + 1}: {row_values}")
                if row_idx >= 4: # Read up to 5 rows
                    break
            
            if sheet_name == '25年4月考勤（4.1-4.30）':
                print("\n  --- All values in Column B (Name Column) for '25年4月考勤（4.1-4.30）' ---")
                for r_idx in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=r_idx, column=2).value # Column B is 1-indexed
                    if cell_value is not None and str(cell_value).strip() != '':
                        print(f"    Row {r_idx}: {cell_value}")
            
            if sheet_name == '上下班打卡_月报':
                print("\n  --- Unique values in columns AV-BU for '上下班打卡_月报' ---")
                unique_values = set()
                # Data rows start from row 5 (Python index 4)
                for row_idx in range(5, sheet.max_row + 1):
                    for col in range(48, 74): # AV (48) to BU (73)
                        cell_value = sheet.cell(row=row_idx, column=col).value
                        if cell_value is not None:
                            cleaned_value = str(cell_value).strip().rstrip(';')
                            cleaned_value = cleaned_value.strip()
                            unique_values.add(cleaned_value)
                print(f"    Unique values: {sorted(list(unique_values))}")

    except Exception as e:
        print(f"Error inspecting {file_path}: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Inspect Excel files")
    parser.add_argument("file_path", help="Path to the Excel file to inspect")
    args = parser.parse_args()

    inspect_excel_file(args.file_path)
