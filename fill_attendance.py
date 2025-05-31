import openpyxl
import os

def fill_attendance_data(target_file, source_file, target_sheet_name, source_sheet_name):
    print(f"--- Starting data transfer ---")
    print(f"Target file: {target_file}, Sheet: {target_sheet_name}")
    print(f"Source file: {source_file}, Sheet: {source_sheet_name}")

    if not os.path.exists(target_file):
        print(f"Error: Target file not found at {target_file}")
        return
    if not os.path.exists(source_file):
        print(f"Error: Source file not found at {source_file}")
        return

    try:
        # Load workbooks
        target_wb = openpyxl.load_workbook(target_file)
        source_wb = openpyxl.load_workbook(source_file)

        # Select sheets
        target_sheet = target_wb[target_sheet_name]
        source_sheet = source_wb[source_sheet_name]

        # Build a dictionary of source data by name
        # Source: Name in A (index 0), data from AV (index 47) to BU (index 72)
        # Data rows start from row 5 (Python index 4)
        source_data_by_name = {}
        for row_idx in range(5, source_sheet.max_row + 1): # Iterate from row 5 to max row
            name = source_sheet.cell(row=row_idx, column=1).value # Column A is 1-indexed
            if name:
                name = str(name).strip() # Ensure string and strip whitespace
                # Extract data from AV (48) to BU (73)
                # openpyxl cell.column is 1-indexed, so AV is 48, BU is 73
                # Python slice is [start_index:end_index], where end_index is exclusive
                # So, for columns 48 to 73 (inclusive), it's cells[47:73]
                row_values = [source_sheet.cell(row=row_idx, column=col).value for col in range(48, 74)]
                source_data_by_name[name] = row_values
        print(f"Loaded {len(source_data_by_name)} entries from source file. Sample names: {list(source_data_by_name.keys())[:5]}")

        # Fill data in target workbook
        # Target: Name in B (index 1), data to A (index 0) to Z (index 25)
        # Data rows start from row 5 (Python index 4)
        # Find the actual last row with a name in the target sheet
        last_name_row = 0
        for r_idx in range(5, target_sheet.max_row + 1):
            name_cell_value = target_sheet.cell(row=r_idx, column=2).value # Column B is 1-indexed
            if name_cell_value is None or str(name_cell_value).strip() == '':
                last_name_row = r_idx - 1 # The previous row was the last one with a name
                break
            last_name_row = r_idx # If loop finishes, max_row is the last name row

        if last_name_row < 5: # Ensure we have at least one data row
            print("No names found in target sheet or data starts after row 4.")
            return

        filled_count = 0
        unmatched_names = []
        for row_idx in range(5, last_name_row + 1): # Iterate from row 5 up to the last name row
            name = target_sheet.cell(row=row_idx, column=2).value # Column B is 1-indexed
            if name:
                name = str(name).strip() # Ensure string and strip whitespace
                if name in source_data_by_name:
                    source_values = source_data_by_name[name]
                    
                    # Define the mapping for text to symbols
                    # Note: These symbols are based on the user's examples.
                    # More mappings might be needed based on other attendance statuses.
                    # Define the mapping for text to symbols based on Row 1 and Row 2 of target file
                    symbol_map = {
                        "正常": "√", # User's example, and '日班' is √
                        "正常（休息）": "T", # User's example, and '休息' is T
                        "正常（未排班）": "T", # User's example
                        "日班": "√",
                        "中班": "@",
                        "晚班": "*",
                        "年休": "N",
                        "调补休": "ィ",
                        "公出": "⊕",
                        "工伤": "±",
                        "婚假": "H",
                        "丧假": "S",
                        "产假": "♀",
                        "产检假": "J",
                        "哺乳假": "P",
                        "加班": "+",
                        "事假": "O",
                        "病假": "△",
                        "旷工": "×",
                        "迟到/上午未打卡": "L", # Base for '迟到'
                        "早退/下午未打卡": "Z", # Base for '早退'
                        "休息": "T",
                        "缺卡": "Q",
                        "育儿假": "Y",
                        "疗休养": "A",
                        "陪产假": "♂",
                        "独生子女陪护假": "D",
                        "居家办公": "G",
                        # Need to confirm for "地点异常" and "--"
                    }

                    # Helper function to get symbol for a status string, handling combinations
                    def get_symbol_for_status(status_text):
                        if not isinstance(status_text, str):
                            return status_text # Return as is if not a string

                        cleaned_text = status_text.strip().rstrip(';').strip()

                        # Check for exact matches first
                        if cleaned_text in symbol_map:
                            return symbol_map[cleaned_text]
                        
                        # Handle combinations or variations
                        found_symbols = []
                        # Order matters for parsing, e.g., "迟到" before "迟到/上午未打卡"
                        # But here we are looking for keywords within the string
                        keywords = {
                            "迟到": "L",
                            "早退": "Z",
                            "缺卡": "Q",
                            "旷工": "×",
                            # "地点异常" and "--" are to be retained as per user, so no symbol mapping here
                        }
                        
                        # Specific handling for "正常" variations that might not be exact matches
                        if "正常（休息）" in cleaned_text:
                            found_symbols.append(symbol_map["正常（休息）"])
                        elif "正常（未排班）" in cleaned_text:
                            found_symbols.append(symbol_map["正常（未排班）"])
                        elif "正常" in cleaned_text:
                            found_symbols.append(symbol_map["正常"])

                        # Check for other keywords
                        if "迟到" in cleaned_text and "L" not in found_symbols: # Avoid duplicates if already added by exact match
                            found_symbols.append(keywords["迟到"])
                        if "早退" in cleaned_text and "Z" not in found_symbols:
                            found_symbols.append(keywords["早退"])
                        if "缺卡" in cleaned_text and "Q" not in found_symbols:
                            found_symbols.append(keywords["缺卡"])
                        if "旷工" in cleaned_text and "×" not in found_symbols:
                            found_symbols.append(keywords["旷工"])
                        
                        # "地点异常" and "--" are handled by the default return cleaned_text

                        # For other specific leaves/statuses that might appear in combinations
                        # This part needs to be more robust if combinations are complex
                        # For now, rely on exact matches for single statuses and keyword search for combinations
                        
                        if found_symbols:
                            return "".join(sorted(list(set(found_symbols)))) # Sort and unique to ensure consistent order
                        
                        # If no specific mapping or keyword found, return original cleaned value
                        return cleaned_text

                    # Copy source_values (26 columns) to target columns C-AB (indices 2-27)
                    for col_offset in range(26): # 0 to 25
                        original_value = source_values[col_offset]
                        processed_value = get_symbol_for_status(original_value)
                        target_sheet.cell(row=row_idx, column=3 + col_offset).value = processed_value
                    filled_count += 1
                else:
                    unmatched_names.append(name)
        print(f"Filled data for {filled_count} names in target file.")
        if unmatched_names:
            print(f"Unmatched names in target file (first 5): {unmatched_names[:5]}")

        # Save the modified target workbook
        target_wb.save(target_file)
        print(f"Successfully updated {target_file}")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    current_directory = os.getcwd()
    target_excel_file = os.path.join(current_directory, "2025年5月考勤.xlsx")
    source_excel_file = os.path.join(current_directory, "上下班打卡_月报_20250501-20250526.xlsx")

    target_sheet = '25年4月考勤（4.1-4.30）'
    source_sheet = '上下班打卡_月报'

    fill_attendance_data(target_excel_file, source_excel_file, target_sheet, source_sheet)
